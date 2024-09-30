import openpyxl
import logging
from datetime import datetime
logging.basicConfig(level=logging.INFO)
def get_current_datetime():
    now = datetime.now()
    formatted_datetime = now.strftime("%m%d-%H%M%S")
    return formatted_datetime


def process(filepath:str,outputPath=None)->str:
    errorLog=""

    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    rowNumber = ws.max_row
    keywordTable = {}
    keywordRow = ws[1]
    for index in range(len(keywordRow)):
        keywordTable[keywordRow[index].value] = index + 1

    # 首先初始化所有计算后价值为单独价值
    ## 当然首先要开一个新col
    calculatedValueIndex = ws.max_column + 1  # 最后新开一列
    if "合并后资产价值" in keywordTable:
        # 如果已经存在的话就覆盖重新算一次
        calculatedValueIndex = keywordTable["合并后资产价值"]
    ws.cell(row=1, column=calculatedValueIndex, value="合并后资产价值")
    for i in range(2, rowNumber+1):
        myValue = ws.cell(row=i, column=keywordTable["单价/元"]).value
        ws.cell(row=i, column=calculatedValueIndex, value=myValue)

    parentAssetSet = set()
    parentAssetTable = {}  # 为了更快地合并自资产的价值，建立一个父资产row_index的字典
    parentIndex = keywordTable["父资产编号"]
    AssetIndex = keywordTable["资产编号"]
    # 这个循环找到所有的负资产
    for i in range(2, rowNumber+1):
        cellValue = str(ws.cell(row=i, column=parentIndex).value)
        if cellValue not in parentAssetSet:
            parentAssetSet.add(cellValue)
    logging.info(f"搜集到{len(parentAssetSet)}个父资产")

    # 记录所有负资产的行号
    for i in range(2, rowNumber+1):
        cellValue: str = str(ws.cell(row=i, column=AssetIndex).value)
        # 如果是-1这种的要把后缀全部忽略掉
        # if "-" in cellValue:
        #     cellValue=cellValue.split("-")[0]
        if cellValue in parentAssetSet:
            parentAssetTable[cellValue] = i
    ## 然后找到所有「是否为配件/增值」为「是」的配件
    isExtra = keywordTable["是否为配件/增值"]
    for i in range(2, rowNumber+1):
        if ws.cell(row=i, column=isExtra).value == "是":
            parentAsset = str(ws.cell(row=i, column=parentIndex).value)  # 我的父资产编号
            if parentAsset not in parentAssetTable:
                errorLog+=f'子资产没有找到对应的父资产{parentAsset}!!\n'
                logging.error(f'子资产没有找到对应的父资产{parentAsset}!!')
                continue
            mergedValue = ws.cell(row=parentAssetTable[parentAsset], column=calculatedValueIndex).value + ws.cell(row=i,column=keywordTable["单价/元"]).value
            ws.cell(row=parentAssetTable[parentAsset], column=calculatedValueIndex, value=mergedValue)
            ws.cell(row=i, column=calculatedValueIndex, value=0)
    if outputPath:
        wb.save(outputPath)
    return errorLog

if __name__ == '__main__':
    process(r"D:\21321.xlsx")