import logging
from datetime import datetime

import openpyxl
from tqdm import tqdm

logging.basicConfig(level=logging.INFO)


def get_current_datetime():
    """
    获取当前的日期和时间。

    返回:
        formatted_datetime (str): 格式化的当前日期和时间，格式为月日-时分秒。
    """
    now = datetime.now()  # 获取当前的日期和时间
    formatted_datetime = now.strftime("%m%d-%H%M%S")  # 将日期和时间格式化为字符串
    return formatted_datetime


def process(filepath: str, outputPath=None) -> str:
    logging.info(f"处理{filepath}, 导出到{outputPath}")
    errorLog = ""

    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    logging.info("已加载表格")

    # Create a mapping from column headers to their indices
    keywordTable = {}
    rowNumber = ws.max_row
    keywordRow = ws[1]
    for index in range(len(keywordRow)):
        keywordTable[keywordRow[index].value] = index + 1
    logging.info("已读取表头")

    # Check if "适用范围" exists in the headers
    if "适用范围" not in keywordTable:
        errorLog = f'Error: "适用范围" column not found in the provided Excel file.'
        return errorLog


    # 记录父资产编号的行号映射， 这样在找父资产的时候能快一点
    asset_number_to_row = {ws.cell(row=i, column=keywordTable["资产编号"]).value: i for i in
                           range(2, rowNumber + 1)}

    logging.info("已建立资产编号哈希表")

    # 遍历列 keywordTable["适用范围"]
    applicable_range_column = [str(cell[0]) for cell in
                               ws.iter_rows(min_row=2, max_row=rowNumber, min_col=keywordTable["适用范围"],
                                            max_col=keywordTable["适用范围"], values_only=True)]
    is_peijian_column = [cell[0] for cell in
                         ws.iter_rows(min_row=2, max_row=rowNumber, min_col=keywordTable["是否为配件/增值"],
                                      max_col=keywordTable["是否为配件/增值"], values_only=True)]
    parent_asset_number_column = [str(cell[0]) for cell in
                                  ws.iter_rows(min_row=2, max_row=rowNumber, min_col=keywordTable["父资产编号"],
                                               max_col=keywordTable["父资产编号"], values_only=True)]
    logging.info("已读取最重要的三列数据")

    copyMap = {}

    def filterOne(filter):
        # 创建一个sheet，名称为filter
        filtered_sheet = wb.create_sheet(title=filter)
        # Copy the header row to the new sheet
        for col in range(1, ws.max_column + 1):
            filtered_sheet.cell(row=1, column=col, value=ws.cell(row=1, column=col).value)

        # 过滤和复制数据
        filtered_row_index = 1  # 记录新sheet的行号
        for i in tqdm(range(rowNumber - 1), desc=f'Filtering for {filter}'):
            cellValue = applicable_range_column[i]
            isPeijian = is_peijian_column[i]
            filterOut = False

            if filter in cellValue and isPeijian == "否":
                filterOut = True

            if isPeijian == "是":
                父资产编号 = parent_asset_number_column[i]
                parent_row = asset_number_to_row.get(父资产编号)
                if parent_row:
                    parent_range_value = applicable_range_column[parent_row]
                    if filter in parent_range_value:
                        filterOut = True

            if filterOut:
                # 复制该行到新sheet中
                for col in range(1, ws.max_column + 1):
                    filtered_sheet.cell(row=filtered_row_index, column=col, value=ws.cell(row=i + 2, column=col).value)
                filtered_row_index += 1


    # 遍历「适用范围」，获取一个set集合
    filterSet = set()
    for i in tqdm(range(2, rowNumber + 1), desc="Collecting filters"):
        cellValue = str(ws.cell(row=i, column=keywordTable["适用范围"]).value)
        isPeijian = ws.cell(row=i, column=keywordTable["是否为配件/增值"]).value
        if isPeijian == "否":
            filterSet.add(cellValue)
    print(filterSet)

    copyMap={filter:[] for filter in filterSet}

    for filter in tqdm(filterSet, desc="Processing filters"):

        filterOne(filter)

    if outputPath:
        wb.save(outputPath)
    else:
        wb.save(filepath)

    return errorLog



if __name__ == '__main__':
    process(filepath=r"v2/data/1000元以上历年设备对平.xlsx",outputPath="v2/data/1000元以上历年设备对平_all_1.xlsx")